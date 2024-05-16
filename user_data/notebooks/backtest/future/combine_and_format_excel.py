import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, Font
from openpyxl import Workbook

# 定义时间周期和趋势
periods = ["1h", "4h", "5m", "15m"]
trends = ["uptrend", "sideways", "downtrend"]

# 定义文件和输出目录
base_dir = "./result_overview"
OUTPUT_DIR = "result_overview"
OUTPUT_FILE_PREFIX = "result_overview"
excel_output_path = os.path.join(OUTPUT_DIR, f"{OUTPUT_FILE_PREFIX}_combined.xlsx")

# 创建输出目录如果不存在
if not os.path.exists(OUTPUT_DIR):
    os.makedirs(OUTPUT_DIR)

# 创建一个新的工作簿
wb = Workbook()
wb.remove(wb.active)  # 删除默认创建的第一个空白工作表

# 合并每个趋势的数据并添加到工作簿中
for trend in trends:
    combined_df = pd.DataFrame()
    
    # 读取每个时间周期的文件并添加时间周期
    for period in periods:
        file = os.path.join(base_dir, f"result_overview_{trend}_{period}.csv")
        if os.path.exists(file):
            df = pd.read_csv(file)
            df.iloc[:, 0] = df.iloc[:, 0] + f"_{period}"
            combined_df = pd.concat([combined_df, df], ignore_index=True)
        else:
            print(f"文件 {file} 不存在，跳过。")
    
    # 将合并后的数据框添加到新的工作表中
    ws = wb.create_sheet(title=trend.capitalize())
    for r in dataframe_to_rows(combined_df, index=False, header=True):
        ws.append(r)
    
    # 设置列宽
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # 获取列字母
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # 设置字体和对齐方式
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = Font(name='Calibri', size=11)

    # 自动调整行高
    for row in ws.iter_rows():
        max_height = 0
        for cell in row:
            if isinstance(cell.value, str):
                cell_height = cell.value.count('\n') + 1
                if cell_height > max_height:
                    max_height = cell_height
        ws.row_dimensions[row[0].row].height = 15 * max_height

# 保存工作簿
wb.save(excel_output_path)
print(f"文件合并和格式化完成，结果已保存到 {excel_output_path}")
